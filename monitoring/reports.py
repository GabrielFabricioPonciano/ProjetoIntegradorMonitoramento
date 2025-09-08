"""
Sistema de geração de relatórios para o monitoramento ambiental.
Suporta exportação em PDF e Excel com gráficos e estatísticas.
"""

import io
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Avg, Min, Max, Count
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.legends import Legend
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Measurement
from .domain import TEMP_LOW, TEMP_HIGH, RH_LIMIT


class ReportGenerator:
    """Gerador de relatórios em PDF e Excel"""
    
    def __init__(self, start_date=None, end_date=None, days=None):
        """
        Inicializa o gerador com período específico.
        
        Args:
            start_date: Data inicial (datetime)
            end_date: Data final (datetime)  
            days: Últimos N dias (int)
        """
        self.timezone_br = timezone.get_current_timezone()
        
        if days:
            # Usar dados mais recentes como referência
            try:
                latest = Measurement.objects.latest('ts')
                self.end_date = latest.ts
                self.start_date = self.end_date - timedelta(days=days)
            except Measurement.DoesNotExist:
                self.end_date = timezone.now()
                self.start_date = self.end_date - timedelta(days=days)
        else:
            self.start_date = start_date or timezone.now() - timedelta(days=30)
            self.end_date = end_date or timezone.now()
            
        self.queryset = Measurement.objects.filter(
            ts__gte=self.start_date,
            ts__lte=self.end_date
        ).order_by('ts')
        
    def get_statistics(self):
        """Calcula estatísticas do período"""
        agg = self.queryset.aggregate(
            temp_avg=Avg('temp_current'),
            temp_min=Min('temp_current'),
            temp_max=Max('temp_current'),
            rh_avg=Avg('rh_current'),
            rh_min=Min('rh_current'),
            rh_max=Max('rh_current'),
            total_count=Count('id')
        )
        
        # Calcular violações
        violations = []
        for measurement in self.queryset:
            temp = measurement.temp_current or 0
            rh = (measurement.rh_current or 0) * 100
            
            if temp < TEMP_LOW or temp > TEMP_HIGH:
                violations.append(measurement)
            elif rh > RH_LIMIT:
                violations.append(measurement)
        
        return {
            'period': {
                'start': timezone.localtime(self.start_date),
                'end': timezone.localtime(self.end_date),
                'days': (self.end_date - self.start_date).days
            },
            'temperature': {
                'avg': round(agg['temp_avg'] or 0, 2),
                'min': round(agg['temp_min'] or 0, 1),
                'max': round(agg['temp_max'] or 0, 1)
            },
            'humidity': {
                'avg': round((agg['rh_avg'] or 0) * 100, 2),
                'min': round((agg['rh_min'] or 0) * 100, 1),
                'max': round((agg['rh_max'] or 0) * 100, 1)
            },
            'measurements': {
                'total': agg['total_count'] or 0,
                'violations': len(violations),
                'compliance': round((1 - len(violations) / max(agg['total_count'], 1)) * 100, 1)
            }
        }
    
    def generate_pdf_report(self):
        """Gera relatório em PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph("Relatório de Monitoramento Ambiental", title_style))
        
        # Estatísticas
        stats = self.get_statistics()
        
        # Informações do período
        period_data = [
            ['Período de Análise', ''],
            ['Data Inicial', stats['period']['start'].strftime('%d/%m/%Y %H:%M')],
            ['Data Final', stats['period']['end'].strftime('%d/%m/%Y %H:%M')],
            ['Duração', f"{stats['period']['days']} dias"],
            ['Total de Medições', str(stats['measurements']['total'])],
        ]
        
        period_table = Table(period_data, colWidths=[3*inch, 2*inch])
        period_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(period_table)
        elements.append(Spacer(1, 20))
        
        # Estatísticas de Temperatura
        temp_data = [
            ['Estatísticas de Temperatura', ''],
            ['Média', f"{stats['temperature']['avg']}°C"],
            ['Mínima', f"{stats['temperature']['min']}°C"],
            ['Máxima', f"{stats['temperature']['max']}°C"],
            ['Limite Inferior', f"{TEMP_LOW}°C"],
            ['Limite Superior', f"{TEMP_HIGH}°C"],
        ]
        
        temp_table = Table(temp_data, colWidths=[3*inch, 2*inch])
        temp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.red),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(temp_table)
        elements.append(Spacer(1, 20))
        
        # Estatísticas de Umidade
        humidity_data = [
            ['Estatísticas de Umidade', ''],
            ['Média', f"{stats['humidity']['avg']}%"],
            ['Mínima', f"{stats['humidity']['min']}%"],
            ['Máxima', f"{stats['humidity']['max']}%"],
            ['Limite Máximo', f"{RH_LIMIT}%"],
        ]
        
        humidity_table = Table(humidity_data, colWidths=[3*inch, 2*inch])
        humidity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightcyan),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(humidity_table)
        elements.append(Spacer(1, 20))
        
        # Análise de Conformidade
        compliance_data = [
            ['Análise de Conformidade', ''],
            ['Total de Violações', str(stats['measurements']['violations'])],
            ['Taxa de Conformidade', f"{stats['measurements']['compliance']}%"],
            ['Status', 'CONFORME' if stats['measurements']['compliance'] >= 95 else 'NÃO CONFORME'],
        ]
        
        compliance_color = colors.green if stats['measurements']['compliance'] >= 95 else colors.orange
        compliance_table = Table(compliance_data, colWidths=[3*inch, 2*inch])
        compliance_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), compliance_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightyellow),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(compliance_table)
        
        # Rodapé
        elements.append(Spacer(1, 40))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1
        )
        elements.append(Paragraph(
            f"Relatório gerado em {timezone.localtime(timezone.now()).strftime('%d/%m/%Y às %H:%M')}<br/>"
            "Sistema de Monitoramento Ambiental - Projeto Integrador IV", 
            footer_style
        ))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_excel_report(self):
        """Gera relatório em Excel com dados e gráfico"""
        workbook = openpyxl.Workbook()
        
        # Planilha de estatísticas
        stats_sheet = workbook.active
        stats_sheet.title = "Estatísticas"
        
        stats = self.get_statistics()
        
        # Cabeçalho
        stats_sheet['A1'] = "RELATÓRIO DE MONITORAMENTO AMBIENTAL"
        stats_sheet['A1'].font = Font(bold=True, size=16)
        stats_sheet.merge_cells('A1:B1')
        
        # Período
        row = 3
        stats_sheet[f'A{row}'] = "PERÍODO DE ANÁLISE"
        stats_sheet[f'A{row}'].font = Font(bold=True)
        row += 1
        stats_sheet[f'A{row}'] = "Data Inicial:"
        stats_sheet[f'B{row}'] = stats['period']['start'].strftime('%d/%m/%Y %H:%M')
        row += 1
        stats_sheet[f'A{row}'] = "Data Final:"
        stats_sheet[f'B{row}'] = stats['period']['end'].strftime('%d/%m/%Y %H:%M')
        row += 1
        stats_sheet[f'A{row}'] = "Duração:"
        stats_sheet[f'B{row}'] = f"{stats['period']['days']} dias"
        
        # Temperatura
        row += 3
        stats_sheet[f'A{row}'] = "TEMPERATURA"
        stats_sheet[f'A{row}'].font = Font(bold=True)
        stats_sheet[f'A{row}'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        row += 1
        stats_sheet[f'A{row}'] = "Média:"
        stats_sheet[f'B{row}'] = f"{stats['temperature']['avg']}°C"
        row += 1
        stats_sheet[f'A{row}'] = "Mínima:"
        stats_sheet[f'B{row}'] = f"{stats['temperature']['min']}°C"
        row += 1
        stats_sheet[f'A{row}'] = "Máxima:"
        stats_sheet[f'B{row}'] = f"{stats['temperature']['max']}°C"
        
        # Umidade
        row += 3
        stats_sheet[f'A{row}'] = "UMIDADE RELATIVA"
        stats_sheet[f'A{row}'].font = Font(bold=True)
        stats_sheet[f'A{row}'].fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        row += 1
        stats_sheet[f'A{row}'] = "Média:"
        stats_sheet[f'B{row}'] = f"{stats['humidity']['avg']}%"
        row += 1
        stats_sheet[f'A{row}'] = "Mínima:"
        stats_sheet[f'B{row}'] = f"{stats['humidity']['min']}%"
        row += 1
        stats_sheet[f'A{row}'] = "Máxima:"
        stats_sheet[f'B{row}'] = f"{stats['humidity']['max']}%"
        
        # Conformidade
        row += 3
        stats_sheet[f'A{row}'] = "CONFORMIDADE"
        stats_sheet[f'A{row}'].font = Font(bold=True)
        color = "4CAF50" if stats['measurements']['compliance'] >= 95 else "FF9800"
        stats_sheet[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1
        stats_sheet[f'A{row}'] = "Total de Medições:"
        stats_sheet[f'B{row}'] = stats['measurements']['total']
        row += 1
        stats_sheet[f'A{row}'] = "Violações:"
        stats_sheet[f'B{row}'] = stats['measurements']['violations']
        row += 1
        stats_sheet[f'A{row}'] = "Taxa de Conformidade:"
        stats_sheet[f'B{row}'] = f"{stats['measurements']['compliance']}%"
        
        # Planilha de dados
        data_sheet = workbook.create_sheet("Dados")
        headers = ['Data/Hora', 'Temperatura (°C)', 'Umidade (%)']
        for col, header in enumerate(headers, 1):
            cell = data_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Dados das medições
        for row, measurement in enumerate(self.queryset[:500], 2):  # Limite 500 registros
            data_sheet.cell(row=row, column=1, value=timezone.localtime(measurement.ts).strftime('%d/%m/%Y %H:%M'))
            data_sheet.cell(row=row, column=2, value=measurement.temp_current or 0)
            data_sheet.cell(row=row, column=3, value=(measurement.rh_current or 0) * 100)
        
        # Gráfico (se houver dados)
        if self.queryset.exists():
            chart = LineChart()
            chart.title = "Monitoramento Ambiental"
            chart.style = 13
            chart.x_axis.title = 'Data/Hora'
            chart.y_axis.title = 'Valores'
            
            # Dados para o gráfico (máximo 50 pontos)
            data_count = min(self.queryset.count(), 50)
            step = max(1, self.queryset.count() // data_count)
            
            temp_data = Reference(data_sheet, min_col=2, min_row=1, max_row=data_count+1)
            humidity_data = Reference(data_sheet, min_col=3, min_row=1, max_row=data_count+1)
            categories = Reference(data_sheet, min_col=1, min_row=2, max_row=data_count+1)
            
            chart.add_data(temp_data, titles_from_data=True)
            chart.add_data(humidity_data, titles_from_data=True)
            chart.set_categories(categories)
            
            data_sheet.add_chart(chart, "E2")
        
        # Ajustar largura das colunas
        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def export_pdf_report(request):
    """View para exportar relatório em PDF"""
    days = int(request.GET.get('days', 30))
    
    generator = ReportGenerator(days=days)
    pdf_content = generator.generate_pdf_report()
    
    response = HttpResponse(pdf_content, content_type='application/pdf')
    filename = f"relatorio_monitoramento_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_excel_report(request):
    """View para exportar relatório em Excel"""
    days = int(request.GET.get('days', 30))
    
    generator = ReportGenerator(days=days)
    excel_content = generator.generate_excel_report()
    
    response = HttpResponse(
        excel_content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"relatorio_monitoramento_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
